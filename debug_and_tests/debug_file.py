import openpyxl
from openpyxl.workbook import Workbook

from pages.data_methods import DataMethods
from pages.xlsx_resulting_table_methods import ResultingTableMethods


def test_debug():
    tab = [['Медкомплекс АВК', 'Термостат НТ-120', 25860],
           ['СКТБ', 'ГП-20 СПУ', 33060],
           ['СКТБ', 'ГП-10 СПУ', 33560],
           ['Медкомплекс АВК', 'ГП-10 СПУ', 20250],
           ['Медкомплекс АВК', 'ГП-20 СПУ', 31250],
           ['НВ-Лаб', 'ГП-10 СПУ', 29680],
           ['НВ-Лаб', 'Термостат НТ-120', 25860]]

    for row in tab:
        if row[0] == 'НВ-Лаб':
            print(row[1], ' стоит ', row[2], ' руб.')


def test_sort_list_last() -> list:
    data_from_pages = [[111, 'автоклав', 'Главмедпроект', "BES-22L-B-LCD", 110000,
                        'https://glavmedproject.ru/product/bes-22l-b-lcd-youjoy-bes-22l-b-lcd-youjoy/'],
                       [75441, 'автоклав', 'НВ-Лаб', 'BES-22L-B-LCD', 160000,
                        'https://www.nv-lab.ru/catalog_info.php?ID=2881'],
                       [7415411, 'автоклав', 'АйТиСтом', 'BES-22L-B-LCD', 200000,
                        'https://itstom.ru/sterilizaciya/avtoklavy/avtoklav-bes-22l-b-lcd/'],
                       [7415411, 'фотометр', 'НВ-Лаб', 'КФК-3', -100500,
                        'https://www.nv-lab.ru/catalog'],
                       [7415411, 'фотометр', 'АйТиСтом', 'КФК-3', 230000,
                        'https://itstom.ru/sterilizaciya/avtoklavy/avtoklav-bes-22l-b-lcd/'],
                       [7415411, 'фотометр', 'Промэколаб', 'КФК-3', 190000,
                        'https://itstom.ru/sterilizaciya/avtoklavy/avtoklav-bes-22l-b-lcd/'],
                       [7415411, 'фотометр', 'Промэколаб', 'КФК-3', 32333,
                        'https://itstom.ru/sterilizaciya/avtoklavy/avtoklav-bes-22l-b-lcd/'],
                       [7415411, 'Стерилизатор', 'АйТиСтом', 'ГК-25', 230000,
                        'https://itstom.ru/sterilizaciya/avtoklavy/avtoklav-bes-22l-b-lcd/'],
                       [7415411, 'Стерилизатор', 'НВ-Лаб', 'ГК-25', 190000,
                        'https://itstom.ru/sterilizaciya/avtoklavy/avtoklav-bes-22l-b-lcd/'],
                       ]

    data_from_pages.sort(key=lambda sort_column: sort_column[1])
    for row in data_from_pages:
        print(row)

    data_temp = data_from_pages
    resulting_data = []

    while data_temp != []:
        value = data_temp[0][1]


        value_qty = DataMethods.get_number_of_items_in_list(value, data_temp)
        temporary_data, data_temp = DataMethods.get_list_of_temporary_data(value_qty, data_temp)
        temporary_data.sort(key=lambda sort_column: sort_column[4])
        target_price = DataMethods.find_target_price(temporary_data)
        resulting_data = DataMethods.get_resulting_data(resulting_data, temporary_data, target_price)

        print(value)
        print(value_qty)
        for row in temporary_data:
            print(row)

        temporary_data.clear()

    final_data = DataMethods.get_final_data(resulting_data)

    print('Печатаем финальные данные: \n')
    for row in final_data:
        print(row)

def test_get_price():
    def change_str_to_num(price_from_site):
        current_price = ''

        for char in price_from_site:
            if char.isdigit() or char == ',' or char == '-' or char == '.':
                current_price = current_price + char

        current_price = current_price.replace(',', '.')
        current_price = current_price.replace('-', '.')

        print('первый символ ', current_price)

        while current_price[0] == '.':
            print(current_price)
            current_price = current_price[1:]

        """while current_price[-1] == '.':
            print(current_price)
            current_price = current_price[:-1]"""



        is_float = False
        while not is_float:
            print('из флот  ', is_float)
            try:
                current_price = float(current_price)
            except ValueError:
                is_float = False
                current_price = current_price[:-1]
                print('Печать из ТРАЯ:   ', current_price)
            else:
                is_float = True


        """if '-00' in current_price:
            current_price = current_price.replace('-00', '')
        if ',00' in current_price:
            current_price = current_price.replace(',00', '')"""
        return current_price


    price_from_site = '239 958 руб. с НДС 20%'
    price = change_str_to_num(price_from_site)

    print('ИТОГ ', price)
    print(2.5+2.5)
    print(2.5 + 2.5)
    print(2.5 + 2.5)
    print(2.5 + 2.5)

def test_find_target_price():
    data_one = [['НВ-Лаб', 'Альтаир 200', 89100],
                ['Экохим', 'Альтаир 200', 127500],
                ['Промэколаб', 'Альтаир 200', 153000],
                ['Промэколаб', 'Альтаир 200', 99800],
                ['Прайм', 'Альтаир 300', 80325]]

    for c in data_one:
        if "НВ-Лаб" in c:
            c.append(1)
            min_price = c[2]
            print(c)
            print('минимальная цена: ', min_price)


def test_get_locator(url='https://www.nv-lab.ru/catalog_info.php?ID=1493'):

    site_title = url.split('://')[-1].split('/')[0]
    locators_dict = openpyxl.load_workbook("select_dict.xlsx")
    locators_dict_sheet = locators_dict.active
    index = 0
    for i in range(1, locators_dict_sheet.max_row + 1):
        site_name = locators_dict_sheet.cell(row=i, column=1).value
        locator = locators_dict_sheet.cell(row=i, column=2).value
        if site_name in site_title:
            site_locator = locator
            index = index + 1
            print(site_locator)