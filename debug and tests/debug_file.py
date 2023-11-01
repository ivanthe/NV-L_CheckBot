import openpyxl
from openpyxl.workbook import Workbook

from pages.data_methods import DataMethods
from pages.xlsx_resulting_table_methods import ResultingTableMethods


def test_debug():
    tab = [['Медкомплекс АВК', 'Термостат НТ-120', 25860],
           ['СКТБ', 'ГП-20 СПУ', 33060],
           ['СКТБ', 'ГП-10 СПУ', 33560],
           ['Медкомплекс АВК', 'ГП-10 СПУ', 20250],
           ['Медкомплекс АВК', 'ГП-20 СПУ', 31250],
           ['НВ-Лаб', 'ГП-10 СПУ', 29680],
           ['НВ-Лаб', 'Термостат НТ-120', 25860]]

    for row in tab:
        if row[0] == 'НВ-Лаб':
            print(row[1], ' стоит ', row[2], ' руб.')

def test_sort():
    data_one = [['НВ-Лаб', 'Альтаир 200', 89100],
                ['Экохим', 'Альтаир 200', 127500],
                ['Промэколаб', 'Альтаир 200', 153000],
                ['Промэколаб', 'Альтаир 200', 99800],
                ['НВ-Лаб', 'Альтаир 300', 127670],
                ['Прайм', 'Альтаир 300', 80325],
                ['Прайм', 'Альтаир 300', 122010],
                ['Экохим', 'Альтаир 300', 185300],
                ['ЗОМЗ', 'Альтаир 300', 129680],
                ['Электронприбор', 'Альтаир 300', 220000],
                ['Электронприбор', 'Альтаир 300', 187000],
                ['Промэколаб', 'Альтаир 300', 143000],
                ['НВ-Лаб', 'Альтаир 300УФ', 212930],
                ['Прайм', 'Альтаир 300УФ', 154350],
                ['Прайм', 'Альтаир 300УФ', 225225],
                ['Экохим', 'Альтаир 300УФ', 283900],
                ['Электронприбор', 'Альтаир 300УФ', 374000],
                ['Эковью', 'Альтаир 300УФ', 238500],
                ['НОВАПРИБОР', 'Альтаир 300УФ', 176000]]


    process = DataMethods(data_one)
    data = process.sort_list(data_one)
    print("")

    def takeSecond(elem):
        return elem[1]
    def takeThird(elem):
        return elem[2]


    def get_number_of_items_in_list(cell_value, current_data):
        b = 0
        for cell in current_data:
            if cell_value in cell:
                b += 1
        return b

    def get_list_of_tempory_data(list_size, current_data):
        data = []
        for c in range(0, list_size):
            data.append(current_data.pop(0))
        return data, current_data

    data.sort(key=takeSecond)
    for row in data:
        print(row)

    def get_resulting_data(resulting_data, current_data):
        data = resulting_data
        for row in current_data:
            data.append(row)
        return data

    data_temp = data
    resulting_data = []

    while data_temp != []:
        value = data_temp[0][1]
        value_qty = get_number_of_items_in_list(value, data_temp)
        tempory_data, data_temp = get_list_of_tempory_data(value_qty, data_temp)
        tempory_data.sort(key=takeThird)
        resulting_data = get_resulting_data(resulting_data, tempory_data)
        tempory_data.clear()

    print('Финальная таблица')
    for row in resulting_data:
        print(row)

    result_workbook = Workbook()
    result_list = result_workbook.active

    xlsx_process = ResultingTableMethods(result_list)
    xlsx_process.create_first_row_of_resulting_table(result_list, resulting_data)
    xlsx_process.format_the_resulting_table(result_list)

    result_workbook.save('debug.xlsx')

def test_get_price():
    def change_str_to_num(price_from_site):
        current_price = ''

        for char in price_from_site:
            if char.isdigit() or char == ',' or char == '-' or char == '.':
                current_price = current_price + char

        current_price = current_price.replace(',', '.')
        current_price = current_price.replace('-', '.')

        print('первый символ ', current_price)

        while current_price[0] == '.':
            print(current_price)
            current_price = current_price[1:]

        while current_price[-1] == '.':
            print(current_price)
            current_price = current_price[:-1]


        """if '-00' in current_price:
            current_price = current_price.replace('-00', '')
        if ',00' in current_price:
            current_price = current_price.replace(',00', '')"""
        return current_price


    price_from_site = 'цена.....,,,,, 20 000-25 руб.'
    price = change_str_to_num(price_from_site)

    print('ИТОГ ', price)
    print(2.5+2.5)

def test_find_target_price():
    data_one = [['НВ-Лаб', 'Альтаир 200', 89100],
                ['Экохим', 'Альтаир 200', 127500],
                ['Промэколаб', 'Альтаир 200', 153000],
                ['Промэколаб', 'Альтаир 200', 99800],
                ['Прайм', 'Альтаир 300', 80325]]

    for c in data_one:
        if "НВ-Лаб" in c:
            c.append(1)
            min_price = c[2]
            print(c)
            print('минимальная цена: ', min_price)


def test_get_locator(url='https://www.nv-lab.ru/catalog_info.php?ID=1493'):

    site_title = url.split('://')[-1].split('/')[0]
    locators_dict = openpyxl.load_workbook("select_dict.xlsx")
    locators_dict_sheet = locators_dict.active
    index = 0
    for i in range(1, locators_dict_sheet.max_row + 1):
        site_name = locators_dict_sheet.cell(row=i, column=1).value
        locator = locators_dict_sheet.cell(row=i, column=2).value
        if site_name in site_title:
            site_locator = locator
            index = index + 1
            print(site_locator)