from openpyxl.styles import PatternFill, Font, Alignment, Border, Side


class XlxsMethods:

    def __init__(self, xlxs_list):
        self.xlxs_list = xlxs_list


    def format_row(self, style, border_style):

        for row in self.xlxs_list.iter_rows(min_row=1, max_col=4, max_row=1):
            for cell in row:
                cell.fill = PatternFill('solid', fgColor=style.font_color)
                cell.font = Font(size=style.font_size, bold=style.font_bold)

        self.xlxs_list.column_dimensions['A'].width = border_style.cell_width * 1.5
        self.xlxs_list.column_dimensions['B'].width = border_style.cell_width * 2
        self.xlxs_list.column_dimensions['C'].width = border_style.cell_width
        self.xlxs_list.column_dimensions['D'].width = border_style.cell_width * 4
        self.xlxs_list.row_dimensions[1].height = border_style.cell_height

        line = Side(border_style=border_style.cell_border_style, color=border_style.cell_border_color)

        for cell in self.xlxs_list[1]:
            cell.alignment = Alignment(horizontal=border_style.cell_alignment_horizontal,
                                       vertical=border_style.cell_alignment_vertical)
            cell.border = Border(top=line, bottom=line, left=line, right=line)
