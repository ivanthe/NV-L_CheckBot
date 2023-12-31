import random
import time
import openpyxl
from pages.base_page import BasePage
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException, InvalidSelectorException


class GeneralMethods(BasePage):

    def get_price(self, locator) -> str:
        # self.go_to_element(self.element_is_present(locator))
        try:
            price_from_site = self.element_is_present(locator).text
            price = self.change_str_to_num(price_from_site)
        except TimeoutException:
            return 'Error'
        except InvalidSelectorException:
            return 'Error'
        return str(price)

    @staticmethod
    def get_locator(url) -> str:
        site_title = url.split('://')[-1].split('/')[0]
        locators_dict = openpyxl.load_workbook("../locators/select_dict.xlsx")
        locators_dict_sheet = locators_dict.active
        site_locator = ''
        for i in range(1, locators_dict_sheet.max_row):
            site_name = locators_dict_sheet.cell(row=i, column=1).value
            locator = locators_dict_sheet.cell(row=i, column=2).value
            if site_name in site_title:
                site_locator = locator
        return site_locator

    @staticmethod
    def change_str_to_num(price_from_site) -> float:

        try:
            current_price = ''

            for char in price_from_site:
                if char.isdigit() or char == ',' or char == '-' or char == '.':
                    current_price = current_price + char

            current_price = current_price.replace(',', '.')
            current_price = current_price.replace('-', '.')

            while current_price[0] == '.':
                current_price = current_price[1:]

            is_float = False
            while not is_float:
                try:
                    current_price = float(current_price)
                except ValueError:
                    is_float = False
                    current_price = current_price[:-1]
                else:
                    is_float = True
        except Exception:
            return 0

        return current_price

    def get_data(self, datafile, result_data) -> list:
        for i in range(2, datafile.max_row + 1):
            c_code = datafile.cell(row=i, column=1).value
            goods_category = datafile.cell(row=i, column=2).value
            goods_name = datafile.cell(row=i, column=3).value
            company_name = datafile.cell(row=i, column=4).value
            current_url = datafile.cell(row=i, column=5).value
            css_selector_type = datafile.cell(row=i, column=6).value
            if css_selector_type == 'стандарт':
                current_css_selector = self.get_locator(current_url)
            else:
                current_css_selector = css_selector_type

            current_locator = (By.CSS_SELECTOR, current_css_selector)
            self.open(current_url)
            time.sleep(random.randint(3, 5))  # на некоторых сайтах цена подгружается после загрузки нужного элемента
            status_code = self.check_link_status_code(current_url)

            """ Ниже получаем цену с сайта, при этом:
            - если при переходе на сайт статус код 404, то цене на выходе временно присвается значение -100500
            - если локатор не может найти цену, то цене на выходе временно присваиватеся значение -100400
            - если на сайте нет цены, например, "цена по запросу", то на выходе цена будет равна 0
            
            дальше эти значения обрабатываются для формирования ошибок в результирующей таблице"""

            if status_code == 404:
                message = f'Ошибка {status_code} для ссылки {current_url}'
                result_data.append([c_code, goods_category, company_name, goods_name, -100500, message])
                print(message)
                self.write_to_logfile(message)
            else:
                price = self.get_price(current_locator)
                if price == 'Error':
                    message = f'Ошибка подбора CSS Selector\nдля ссылки {current_url}\n или ссылка ' \
                              f'не на карточку товара'
                    result_data.append([c_code, goods_category, company_name, goods_name, -100400, message])
                    print(message)
                    self.write_to_logfile(message)
                elif price == '':
                    result_data.append([c_code, goods_category, company_name, goods_name, 0, current_url])
                else:
                    result_data.append([c_code, goods_category, company_name, goods_name, float(price), current_url])
        return result_data

    @staticmethod
    def write_to_logfile(message) -> None:
        with open('../errorlogs.txt', 'a') as errorlogs:
            errorlogs.write(f'{time.strftime("%m/%d/%Y, %H:%M:%S", time.localtime())}:  {message} \n')
            errorlogs.close()
