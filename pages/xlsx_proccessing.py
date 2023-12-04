from openpyxl.styles import PatternFill, Font, Alignment, Border, Side


class XlsxProccessing:

    def __init__(self, xlxs_list):
        self.xlxs_list = xlxs_list

    @staticmethod
    def format_row(style, border_style, row) -> None:

        line = Side(border_style=border_style.cell_border_style, color=border_style.cell_border_color)

        for cell in row:
            cell.fill = PatternFill('solid', fgColor=style.font_color)
            cell.font = Font(size=style.font_size, bold=style.font_bold)
            cell.alignment = Alignment(horizontal=border_style.cell_alignment_horizontal,
                                       vertical=border_style.cell_alignment_vertical)
            cell.border = Border(top=line, bottom=line, left=line, right=line)

    @staticmethod
    def format_cell(color=None, cell=None) -> None:
        cell.fill = PatternFill('solid', fgColor=color)

