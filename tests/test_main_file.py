from debug_and_tests.debug_data import DebugData
from pages.data_methods import DataMethods
from pages.site_pages import GeneralMethods
import openpyxl
from openpyxl import Workbook
from pages.xlsx_resulting_table_methods import ResultingTableMethods
from tests.check_datafile import CheckLocators


class TestMain:

    def test_get_price(self, driver):
        # Получаем данные из файла datafile и записываем в промежуточный массив
        select_dict = openpyxl.load_workbook('../locators/select_dict.xlsx')
        datafile = openpyxl.load_workbook('../datafile.xlsx')
        select_dict_sheet = select_dict.active
        datafile_sheet = datafile.active
        check_locators = CheckLocators()
        result_check_locators = check_locators.check_if_datafile_locators_in_dictionary(select_dict_sheet,
                                                                                        datafile_sheet)
        if result_check_locators:
            site_processing = GeneralMethods(driver)
            data_from_pages = []

            data_from_pages = site_processing.get_data(datafile_sheet, data_from_pages)

            # Делаем обработку промежуточного массива: сортируем данные

            data_methods = DataMethods(data_from_pages)
            data_from_pages = data_methods.sort_list()

            # Создаем таблицу Ecel с данными для пользователя, форматируем ее и сохраняем

            result_workbook = Workbook()
            result_list = result_workbook.active

            xlsx_process = ResultingTableMethods(result_list)
            xlsx_process.create_first_row_of_resulting_table(result_list, data_from_pages)
            xlsx_process.format_the_resulting_table(result_list)

            result_workbook.save('../debug.xlsx')
        else:
            print('нет подходящих локаторов')

    def test_debug(self, driver):

        """site_processing = GeneralMethods(driver)
        site_processing.open('https://nowsecure.nl')
        # site_processing.open('https://intoli.com/blog/not-possible-to-block-chrome-headless/chrome-headless-test.html')
        time.sleep(25)"""

        print('The debug_function is not use')
