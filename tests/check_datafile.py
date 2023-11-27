import sys

import openpyxl
import xlwt


class CheckLocators:

    @staticmethod
    def check_if_datafile_locators_in_dictionary(select_dict_sheet, datafile_sheet) -> bool:
        data = []
        counter = 0

        for c in range(1, select_dict_sheet.max_row + 1):
            data.append(select_dict_sheet.cell(row=c, column=1).value)

        for i in range(2, datafile_sheet.max_row):
            url = datafile_sheet.cell(row=i, column=5).value
            print('ПОЛУЧИЛИ', url)
            site_title = url.split('://')[-1].split('/')[0]
            print('Это прошли')

            if 'www.' in site_title:
                site_title = site_title.split('www.')[-1]

            if datafile_sheet.cell(row=i, column=6).value is None:
                print(f'Для сайта {site_title} НЕ ПРИСВОЕН НИКАКОЙ ЛОКАТАТОР.')
                counter = counter + 1
            elif datafile_sheet.cell(row=i, column=6).value == 'стандарт':
                if datafile_sheet.cell(row=i, column=6).value == 'стандарт':
                    if site_title not in data:
                        print(f'Для сайта {site_title} нет стандартного локатора в словаре.')
                        counter = counter + 1

        if counter != 0:
            return False
        else:
            return True
