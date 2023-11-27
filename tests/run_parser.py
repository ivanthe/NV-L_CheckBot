import time

import openpyxl

from pages.data_methods import DataMethods
from pages.site_pages import GeneralMethods
from pages.xlsx_resulting_table_methods import ResultingTableMethods
from tests.check_datafile import CheckLocators
from tests.conftest import driver
from openpyxl import Workbook



if __name__ == '__main__':
    select_dict = openpyxl.load_workbook('select_dict.xlsx')
    datafile = openpyxl.load_workbook('datafile.xlsx')
    select_dict_sheet = select_dict.active
    datafile_sheet = datafile.active
    check_locators = CheckLocators()
    check_locators.check_if_datafile_locators_in_dictionary(select_dict_sheet, datafile_sheet)

    site_processing = GeneralMethods(driver)
    time.sleep(5)
    data_from_pages = []

    data_from_pages = site_processing.get_data(datafile_sheet, data_from_pages)

    # Делаем обработку промежуточного массива: сортируем данные

    data_methods = DataMethods(data_from_pages)
    data_from_pages = data_methods.sort_list(data_from_pages)

    # Создаем таблицу Ecel с данными для пользователя, форматируем ее и сохраняем

    result_workbook = Workbook()
    result_list = result_workbook.active

    xlsx_process = ResultingTableMethods(result_list)
    xlsx_process.create_first_row_of_resulting_table(result_list, data_from_pages)
    xlsx_process.format_the_resulting_table(result_list)

    result_workbook.save('debug.xlsx')